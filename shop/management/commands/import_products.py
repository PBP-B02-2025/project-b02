import os
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from shop.models import Product  # Ganti 'shop' jika nama app Anda berbeda
from django.contrib.auth.models import User
import openpyxl  # <-- Import library baru

class Command(BaseCommand):
    help = 'Memasukkan data produk dari file XLSX'

    def handle(self, *args, **kwargs):
        
        # --- Mencari User Default ---
        try:
            default_user = User.objects.filter(is_superuser=True).first()
        except User.DoesNotExist:
             raise CommandError("Tidak ada Superuser (admin) yang ditemukan. Harap buat satu superuser terlebih dahulu.")

        if not default_user:
            raise CommandError("Tidak ada Superuser (admin) yang ditemukan. Harap buat satu superuser terlebih dahulu.")

        self.stdout.write(self.style.SUCCESS(f"Semua produk akan dihubungkan ke user: {default_user.username}"))
        
        # --- Mencari File XLSX ---
        # GANTI NAMA FILE DI SINI JIKA PERLU
        file_name = 'products.xlsx' 
        file_path = os.path.join(settings.BASE_DIR, file_name)
        self.stdout.write(f'Memulai import dari {file_path}...')

        if not os.path.exists(file_path):
             raise CommandError(f"File {file_path} tidak ditemukan. Pastikan file '{file_name}' ada di folder utama proyek Anda (sejajar manage.py).")

        # --- Mulai Proses Impor ---
        try:
            # Buka workbook (file Excel)
            workbook = openpyxl.load_workbook(file_path)
            # Ambil sheet pertama (sheet aktif)
            sheet = workbook.active

            # Buat mapping nama kolom ke indeks (A=1, B=2, dst.)
            # Ini membaca baris PERTAMA (header)
            headers = {}
            for col_index, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value.strip().lower()] = col_index

            # Cek kolom wajib
            required_cols = ['name', 'price', 'size', 'brand', 'description', 'category', 'thumbnail', 'is_featured']
            for col in required_cols:
                if col not in headers:
                    raise CommandError(f"File Excel Anda kekurangan kolom header: '{col}'. Pastikan baris pertama berisi semua nama kolom.")

            total_created = 0
            total_updated = 0
            total_skipped = 0

            # Mulai iterasi dari BARIS KEDUA (iter_rows(min_row=2))
            for row in sheet.iter_rows(min_row=2):
                # Ubah data baris menjadi dictionary agar mudah dibaca
                row_data = {}
                for header_name, col_index in headers.items():
                    # Ambil nilai sel. Indeks openpyxl berbasis 1, tapi row adalah tuple (berbasis 0)
                    row_data[header_name] = row[col_index - 1].value 

                # Ambil data dari Excel.
                # Kita pakai .get() untuk menghindari error jika kolom opsional tidak ada
                product_name = row_data.get('name')
                product_price = row_data.get('price')
                product_size = row_data.get('size')
                product_brand = row_data.get('brand')
                product_desc = row_data.get('description')
                product_category = row_data.get('category')
                product_thumbnail = row_data.get('thumbnail')
                product_is_featured = row_data.get('is_featured')

                # Jika nama produk kosong, lewati baris
                if not product_name:
                    total_skipped += 1
                    continue

                # --- Konversi Tipe Data ---
                try:
                    product_price = int(product_price)
                except (ValueError, TypeError):
                    self.stdout.write(self.style.ERROR(f"Harga tidak valid untuk '{product_name}': {product_price}. Baris dilewati."))
                    total_skipped += 1
                    continue
                
                # Konversi boolean
                is_featured_bool = str(product_is_featured).lower() in ['true', '1', 'yes']

                # --- Simpan ke Database ---
                obj, created = Product.objects.update_or_create(
                    name=product_name,  # Field unik untuk mencari
                    defaults={
                        'user': default_user,
                        'price': product_price,
                        'size': product_size,
                        'brand': product_brand,
                        'description': product_desc,
                        'category': product_category,
                        'thumbnail': product_thumbnail,
                        'is_featured': is_featured_bool,
                    }
                )

                if created:
                    total_created += 1
                else:
                    total_updated += 1
            
            self.stdout.write(self.style.SUCCESS(f'\nImport Selesai!'))
            self.stdout.write(self.style.SUCCESS(f'{total_created} produk baru dibuat.'))
            self.stdout.write(self.style.WARNING(f'{total_updated} produk di-update.'))
            self.stdout.write(self.style.ERROR(f'{total_skipped} baris dilewati (error/kosong).'))

        except Exception as e:
            raise CommandError(f"Terjadi error saat memproses file: {e}")